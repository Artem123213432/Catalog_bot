from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import os
from typing import List, Dict

EXCEL_FILE = "orders.xlsx"

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Заказы"
        
        
        headers = ["ID", "Дата", "Клиент", "Телефон", "Адрес", "Состав заказа", "Сумма"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        
        ws.column_dimensions['A'].width = 10  
        ws.column_dimensions['B'].width = 20  
        ws.column_dimensions['C'].width = 30  
        ws.column_dimensions['D'].width = 20  
        ws.column_dimensions['E'].width = 40  
        ws.column_dimensions['F'].width = 50  
        ws.column_dimensions['G'].width = 15  
        
        wb.save(EXCEL_FILE)

def save_order(order_data: Dict) -> None:
    
    if not os.path.exists(EXCEL_FILE):
        init_excel()
    
    
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    
    last_row = ws.max_row + 1
    
    
    order_items = []
    total_sum = 0
    for item in order_data["items"]:
        quantity = item.get('quantity', 1)
        price = item['price']
        total_sum += price * quantity
        order_items.append(f"{item['name']} x{quantity} - {price} ₽")
    
    
    ws.cell(row=last_row, column=1).value = last_row - 1  
    ws.cell(row=last_row, column=2).value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  
    ws.cell(row=last_row, column=3).value = order_data["client"]["name"]  
    ws.cell(row=last_row, column=4).value = order_data["client"]["phone"]  
    ws.cell(row=last_row, column=5).value = order_data["client"]["address"]  
    ws.cell(row=last_row, column=6).value = "\n".join(order_items)  
    ws.cell(row=last_row, column=7).value = f"{total_sum} ₽"  
    
    
    for col in range(1, 8):
        cell = ws.cell(row=last_row, column=col)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    
    
    wb.save(EXCEL_FILE) 